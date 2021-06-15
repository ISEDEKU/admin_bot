import os
import re

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl.utils import get_column_letter

from data.config import USERS
from loader import dp
from states.state_class import Questions


@dp.message_handler(state=Questions.serial_num, user_id=USERS)  # указываю второе состояние
async def answer_serial_num(message: types.Message, state=FSMContext):
    path_to_file_1 = os.path.abspath('out.xlsx')
    path_to_file_2 = os.path.abspath('stock_balance.xlsx ')
    path_to_file_3 = os.path.abspath('boriychuk.xlsx')
    path_to_file_4 = os.path.abspath('other.xlsx')


    path_to_file = [path_to_file_1, path_to_file_2, path_to_file_3, path_to_file_4]

    data = await state.get_data()
    name_list = data.get('answer1')
    i = data.get('answer2')
    answer = message.text
    fin_mess_graph = {}

    try:
        for path in path_to_file:

            graph = {}

            regular = str(name_list[int(answer)-1])  # в переменную answer помещяется то, что ответил пользователь в ТГ
            wb = openpyxl.load_workbook(path)  # Грузим наш прайс-лист
            sheets_list = wb.sheetnames  # Получаем список всех листов в файле
            sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
            row_max = sheet_active.max_row  # Получаем количество столбцов
            column_max = 3

            row_min = 1  # Переменная, отвечающая за номер строки
            column_min = 1  # Переменная, отвечающая за номер столбца

            while column_min <= column_max:
                row_min_min = row_min
                row_max_max = row_max
                while row_min_min <= row_max_max:
                    row_min_min = str(row_min_min)

                    word_column = get_column_letter(column_min)
                    word_column = str(word_column)
                    word_cell = word_column + row_min_min
                    data_from_cell = sheet_active[word_cell].value
                    data_from_cell = str(data_from_cell)
                    result = re.findall(regular.lower(), data_from_cell.lower())

                    if len(result) > 0:
                        graph['address'] = word_cell
                        graph['path'] = path
                    # каждая ячейка, содержащая в себе нужное нам значение, помещается в лист

                    row_min_min = int(row_min_min)
                    row_min_min = row_min_min + 1
                column_min = column_min + 1

            if len(graph) > 0:

                if graph['path'] == path_to_file_1:

                    sel_num = graph['address'][1:]
                    sel_num = int(sel_num)
                    main_product_name = sheet_active[sel_num][1].value
                    product_name = sheet_active[sel_num][1].value
                    main_manufacture = sheet_active[sel_num][2].value
                    main_quantity = sheet_active[sel_num][3].value
                    cost_opt = sheet_active[sel_num][4].value
                    cost_mem = sheet_active[sel_num][5].value
                    main_cost_dlr = sheet_active[sel_num][6].value
                    main_cost_grn = sheet_active[sel_num][7].value

                    fin_mess_graph['main_quantity'] = f'\nСклад "Навионика" кол-во: {main_quantity}'

                elif graph['path'] == path_to_file_2:
                    sel_num = graph['address'][1:]
                    sel_num = int(sel_num)
                    product_name = sheet_active[sel_num][1].value
                    main_manufacture = sheet_active[sel_num][2].value
                    sh_quantity = sheet_active[sel_num][3].value
                    cost = sheet_active[sel_num][5].value
                    try:
                        sh_opt_cost = sheet_active[sel_num][4].value
                        sh_opt_cost = round(float(sh_opt_cost))
                    except:
                        sh_opt_cost = 'Уточняйте'
                    fin_mess_graph['sh_quantity'] = f'\nСклад "Магазин" кол-во: {sh_quantity}'

                elif graph['path'] == path_to_file_3:
                    sel_num = graph['address'][1:]
                    sel_num = int(sel_num)
                    product_name = sheet_active[sel_num][1].value
                    main_manufacture = sheet_active[sel_num][2].value
                    b_quantity = sheet_active[sel_num][5].value
                    try:
                        cost = float(sheet_active[sel_num][4].value) * 1.1
                        cost = round(cost)
                        b_opt_cost = sheet_active[sel_num][4].value
                    except:
                        cost = 'Уточняйте'
                        b_opt_cost = 'Уточняйте'

                    fin_mess_graph['b_quantity'] = f'\nСклад "Борийчук" кол-во: {b_quantity}'

                elif graph['path'] == path_to_file_4:
                    sel_num = graph['address'][1:]
                    sel_num = int(sel_num)
                    product_name = sheet_active[sel_num][1].value
                    main_manufacture = sheet_active[sel_num][2].value
                    o_quantity = sheet_active[sel_num][5].value
                    try:
                        cost = float(sheet_active[sel_num][4].value) * 1.1
                        cost = round(cost)
                        o_opt_cost = sheet_active[sel_num][4].value
                    except:
                        cost = 'Уточняйте'
                        o_opt_cost = 'Уточняйте'

                    fin_mess_graph['o_quantity'] = f'\nСклад "Другие поставщики" кол-во: {o_quantity}'

        fin_mess = ''
        try:
            fin_mess += f'\n{main_product_name}'
        except:
            fin_mess += f'\n{product_name}'

        fin_mess += f'\nШифр производителя: {main_manufacture}'

        try:
            fin_mess += f'\nРРЦ долар с главного склада: {main_cost_dlr}'
            fin_mess += f'\nРРЦ грн с главного склада: {main_cost_grn}'
            fin_mess += f'\nОпт цена для партнёров: {cost_mem}'
        except:
            fin_mess += f'\nОпт цена для партнёров: {cost}'
        for quantity in fin_mess_graph:
            fin_mess += fin_mess_graph[quantity]
        await message.answer(fin_mess)

    except:
        await message.answer('Такого номера нет в списке...')

