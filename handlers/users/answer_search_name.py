import os
import re
import urllib

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl.utils import get_column_letter

from data.config import ADMINS
from keyboards.default.main_keyboard import menu_1
from loader import dp
from states.state_class import Questions


@dp.message_handler(state=Questions.search_name,
                    user_id=ADMINS)  # хэндлер, с первым состоянием, который принимает ответ
async def answer_search_name(message: types.Message, state=FSMContext):
    await message.answer('Ищу все товары с похожим названием...')
    urllib.request.urlretrieve('http://aparat.ua/cli/telegram_opt.xlsx', 'main.xlsx')
    path_to_file_1 = os.path.abspath('main.xlsx')

    urllib.request.urlretrieve('http://aparat.ua/xlsx/stock_balance.xlsx', 'stock_balance.xlsx ')
    path_to_file_2 = os.path.abspath('stock_balance.xlsx ')

    urllib.request.urlretrieve('http://aparat.ua/xlsx/boriychuk.xlsx', 'boriychuk.xlsx')
    path_to_file_3 = os.path.abspath('boriychuk.xlsx')

    urllib.request.urlretrieve('http://aparat.ua/xlsx/kovel.xlsx', 'kovel.xlsx')
    path_to_file_4 = os.path.abspath('kovel.xlsx')

    urllib.request.urlretrieve('http://garminua.org/cli/garminua.xlsx', 'other.xlsx')
    path_to_file_5 = os.path.abspath('other.xlsx')

    path_to_file = [path_to_file_1, path_to_file_2, path_to_file_3, path_to_file_4, path_to_file_5]
    products = set()

    for path in path_to_file:

        regular = str(message.text)  # в переменную answer помещяется то, что ответил пользователь в ТГ
        wb = openpyxl.load_workbook(path)  # Грузим наш прайс-лист
        sheets_list = wb.sheetnames  # Получаем список всех листов в файле
        sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
        row_max = sheet_active.max_row  # Получаем количество столбцов
        column_max = 3

        row_min = 1  # Переменная, отвечающая за номер строки
        column_min = 1  # Переменная, отвечающая за номер столбца

        while column_min <= column_max:
            row_min_min = row_min
            row_max_max = row_max
            while row_min_min <= row_max_max:
                row_min_min = str(row_min_min)

                word_column = get_column_letter(column_min)
                word_column = str(word_column)
                word_cell = word_column + row_min_min
                data_from_cell = sheet_active[word_cell].value
                data_from_cell = str(data_from_cell)
                result = re.findall(regular.lower(), data_from_cell.lower())

                if len(result) > 0:
                    products.add(str(data_from_cell))
                # каждая ячейка, содержащая в себе нужное нам значение, помещается в лист

                row_min_min = int(row_min_min)
                row_min_min = row_min_min + 1
            column_min = column_min + 1
            # алгоритм для сравнения ответа из ТГ и товарами в exel файле
    if len(products) == 0:
        await message.answer('Оу, я не нашёл ничего подобного...\nНапишите другое название:')
        await state.reset_state()
        await Questions.search_name.set()

    else:
        products = sorted(list(products))
        name_list = []
        if len(products) > 1:
            i = 1
            for product in products:
                name_list.append(f'{product} --- [{i}]')
                i += 1

            try:
                await message.answer(("\n\n".join(name_list)))
                await message.answer(f'Я нашел {len(products)} товаров с похожим названием!\nВведите нужный номер:')
                await state.update_data(answer2=i)
                await state.update_data(answer1=products)
                await Questions.serial_num.set()  # запуск второго состоянияx
            except:
                try:
                    srt = round(len(name_list) / 2)
                    await  message.answer(("\n\n".join(name_list[:srt])))
                    await message.answer(("\n\n".join(name_list[srt:])))
                    await message.answer(f'Я нашел {len(products)} товаров с похожим названием!\nВведите нужный номер:')
                    await state.update_data(answer2=i)
                    await state.update_data(answer1=products)
                    await Questions.serial_num.set()  # запуск второго состоянияx
                except:
                    await message.answer('Оу... Список слишком большой! Будьте поточнее!')



        elif len(products) == 1:
            fin_mess_graph = {}
            for path in path_to_file:
                graph = {}

                regular = str(*products)  # в переменную answer помещяется то, что ответил пользователь в ТГ
                wb = openpyxl.load_workbook(path)  # Грузим наш прайс-лист
                sheets_list = wb.sheetnames  # Получаем список всех листов в файле
                sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
                row_max = sheet_active.max_row  # Получаем количество столбцов
                column_max = sheet_active.max_column  # Получаем количество строк

                row_min = 1  # Переменная, отвечающая за номер строки
                column_min = 3

                while column_min <= column_max:
                    row_min_min = row_min
                    row_max_max = row_max
                    while row_min_min <= row_max_max:
                        row_min_min = str(row_min_min)

                        word_column = get_column_letter(column_min)
                        word_column = str(word_column)
                        word_cell = word_column + row_min_min
                        data_from_cell = sheet_active[word_cell].value
                        data_from_cell = str(data_from_cell)
                        result = re.findall(regular.lower(), data_from_cell.lower())

                        if len(result) > 0:
                            graph['address'] = word_cell
                            graph['path'] = path
                        # каждая ячейка, содержащая в себе нужное нам значение, помещается в лист

                        row_min_min = int(row_min_min)
                        row_min_min = row_min_min + 1
                    column_min = column_min + 1

                if len(graph) > 0:

                    if graph['path'] == path_to_file_1:

                        sel_num = graph['address'][1:]
                        sel_num = int(sel_num)
                        product_name = sheet_active[sel_num][1].value
                        main_manufacture = sheet_active[sel_num][2].value
                        quantity = sheet_active[sel_num][3].value
                        cost_opt = sheet_active[sel_num][4].value
                        cost_mem = sheet_active[sel_num][5].value
                        main_cost_dlr = sheet_active[sel_num][6].value
                        main_cost_grn = sheet_active[sel_num][7].value

                        fin_mess_graph['main_cost_dlr'] = f'\nРРЦ долар с главного склада: {main_cost_dlr}'
                        fin_mess_graph['main_cost_grn'] = f'\nРРЦ грн с главного склада: {main_cost_grn}'

                        await message.answer(
                            f'\n"Основной склад"\n\n{product_name}\nШифр производителя: {main_manufacture}\nКол-во: {quantity} \nОпт $: {cost_opt.lstrip()} \nЦена, партнёры, уе: {cost_mem.lstrip()}\nРРЦ $: {main_cost_dlr.lstrip()}\nРРЦ $ грн: {main_cost_grn.lstrip()}')

                    elif graph['path'] == path_to_file_2:
                        sel_num = graph['address'][1:]
                        sel_num = int(sel_num)
                        product_name = sheet_active[sel_num][1].value
                        main_manufacture = sheet_active[sel_num][2].value
                        quantity = sheet_active[sel_num][3].value
                        cost = sheet_active[sel_num][5].value
                        sh_opt_cost = sheet_active[sel_num][4].value

                        fin_mess_graph['sh_opt_cost'] = f'\nСклад "На магазине" опт цена: {sh_opt_cost}'

                        await message.answer(
                            f'\n"На магазине"\n\n{product_name}\nШифр производителя: {main_manufacture}\nКоличество: {quantity}\nЦена, партнёры, уе: {cost}\nОпт цена в 1с $: {sh_opt_cost}')

                    elif graph['path'] == path_to_file_3:
                        sel_num = graph['address'][1:]
                        sel_num = int(sel_num)
                        product_name = sheet_active[sel_num][1].value
                        main_manufacture = sheet_active[sel_num][2].value
                        quantity = sheet_active[sel_num][5].value
                        cost = float(sheet_active[sel_num][4].value) * 1.1
                        cost = round(cost)
                        b_opt_cost = sheet_active[sel_num][4].value

                        fin_mess_graph['b_opt_cost'] = f'\nСклад "Борийчук" опт цена: {b_opt_cost}'

                        await message.answer(
                            f'\n"Борийчук"\n\n{product_name}\nШифр производителя: {main_manufacture}\nКоличество: {quantity}\nЦена, партнёры, уе: {cost}\nОпт $: {b_opt_cost}')

                    elif graph['path'] == path_to_file_4:
                        sel_num = graph['address'][1:]
                        sel_num = int(sel_num)
                        product_name = sheet_active[sel_num][1].value
                        main_manufacture = sheet_active[sel_num][2].value
                        quantity = sheet_active[sel_num][5].value
                        cost = float(sheet_active[sel_num][4].value) * 1.1
                        cost = round(cost)
                        k_opt_cost = sheet_active[sel_num][4].value

                        fin_mess_graph['k_opt_cost'] = f'\nСклад "Ковель" опт цена: {k_opt_cost}'

                        await message.answer(
                            f'\n"Ковель"\n\n{product_name}\nШифр производителя: {main_manufacture}\nКоличество: {quantity}\nЦена, партнёры, уе: {cost}\nОпт $: {k_opt_cost}')

                    elif graph['path'] == path_to_file_5:
                        sel_num = graph['address'][1:]
                        sel_num = int(sel_num)
                        product_name = sheet_active[sel_num][1].value
                        main_manufacture = sheet_active[sel_num][2].value
                        quantity = sheet_active[sel_num][4].value
                        cost = float(sheet_active[sel_num][4].value) * 1.1
                        cost = round(cost)
                        o_opt_cost = sheet_active[sel_num][4].value

                        fin_mess_graph['o_opt_cost'] = f'\nСклад "Другие поставщики" опт цена: {o_opt_cost}'

                        await message.answer(
                            f'\n"Другие поставщики"\n\n{product_name}\nШифр производителя: {main_manufacture}\nКоличество: {quantity}\nЦена, партнёры, уе: {cost}\nОпт $: {o_opt_cost}')

                fin_mess = f'\n{product_name}\n' + f'\nШифр производителя: {main_manufacture}'
        for cost in fin_mess_graph:
            fin_mess += fin_mess_graph[cost]
        await message.answer(fin_mess)

        await message.answer('Поиск завершён!', reply_markup=menu_1)
        await state.finish()
